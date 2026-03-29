import openpyxl
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from .models import Department, JobPosition, Worker, CurriculumLaboris, WorkerImportLog
from .serializers import (DepartmentSerializer, JobPositionSerializer,
    WorkerListSerializer, WorkerDetailSerializer, CurriculumLaborisSerializer, WorkerImportLogSerializer)


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'code']


class JobPositionViewSet(viewsets.ModelViewSet):
    queryset = JobPosition.objects.select_related('department').all()
    serializer_class = JobPositionSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['department', 'risk_level']
    search_fields = ['title', 'code']


class WorkerViewSet(viewsets.ModelViewSet):
    queryset = Worker.objects.select_related('department', 'job_position').all()
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['department', 'job_position', 'contract_type', 'is_active', 'gender']
    search_fields = ['matricule', 'first_name', 'last_name', 'phone']
    ordering_fields = ['last_name', 'hire_date', 'matricule']

    def get_serializer_class(self):
        if self.action == 'list':
            return WorkerListSerializer
        return WorkerDetailSerializer

    def destroy(self, request, *args, **kwargs):
        worker = self.get_object()
        worker.is_active = False
        worker.save()
        return Response({'detail': 'Travailleur désactivé.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get', 'post'], url_path='curriculum-laboris')
    def curriculum_laboris(self, request, pk=None):
        worker = self.get_object()
        if request.method == 'GET':
            qs = worker.curriculum_laboris.all()
            return Response(CurriculumLaborisSerializer(qs, many=True).data)
        serializer = CurriculumLaborisSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(worker=worker)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser])
    def import_workers(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Fichier requis.'}, status=400)

        log = WorkerImportLog.objects.create(
            imported_by=request.user,
            file_name=file.name,
            status='PROCESSING'
        )
        errors = []
        success = 0
        total = 0
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                total += 1
                data = dict(zip(headers, row))
                try:
                    dept, _ = Department.objects.get_or_create(
                        code=str(data.get('Département', 'N/A')),
                        defaults={'name': str(data.get('Département', 'N/A'))}
                    )
                    pos, _ = JobPosition.objects.get_or_create(
                        code=str(data.get('Poste', 'N/A')),
                        defaults={'title': str(data.get('Poste', 'N/A')), 'department': dept}
                    )
                    Worker.objects.update_or_create(
                        matricule=str(data.get('Matricule', '')),
                        defaults={
                            'first_name': str(data.get('Prénom', '')),
                            'last_name': str(data.get('Nom', '')),
                            'gender': str(data.get('Genre', 'M'))[:1],
                            'department': dept,
                            'job_position': pos,
                            'imported_from_hr': True,
                        }
                    )
                    success += 1
                except Exception as e:
                    errors.append({'row': total, 'error': str(e)})
        except Exception as e:
            log.status = 'FAILED'
            log.errors = [str(e)]
            log.save()
            return Response({'detail': str(e)}, status=500)

        log.row_count = total
        log.success_count = success
        log.error_count = len(errors)
        log.errors = errors
        log.status = 'DONE'
        log.save()
        return Response(WorkerImportLogSerializer(log).data, status=201)
